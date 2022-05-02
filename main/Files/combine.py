# from openpyxl import load_workbook
import datetime
import time

from openpyxl import load_workbook, Workbook
import csv

# first_arr = []  # data from dice
# second_arr = []  # data from apollo
third_arr = []  # Combined data
fourth_arr = []  # unfounded data (people not found for the companies)

day = datetime.datetime.now().strftime("%B-%d")
now = datetime.datetime.now().strftime("%B-%I-%M-%S")

########################################################################################################################
# Blacklist companies and already collected data

workbook99 = load_workbook(filename='Blacklist.xlsx')
workbook98 = load_workbook(filename='QA/Processed 2feb.xlsx')  # Sheet linking to already collected data
sheet99 = workbook99.active
sheet98 = workbook98.active
blacklist = []
column = sheet99.max_row
column2 = sheet98.max_row
for col in range(1, column + 1):
    try:
        valu = sheet99._get_cell(col, 1).value
        valu = valu.lower()
        blacklist.append(valu)
    except:
        pass
for col2 in range(1, column2 + 1):
    try:
        valuz = sheet98._get_cell(col2, 1).value
        valuz = valuz.lower()
        blacklist.append(valuz)
    except:
        pass

########################################################################################################################


########################################################################################################################

# contains data from apollo
workbook1 = load_workbook(filename="All People.xlsx")
sheet1 = workbook1.active
rows = []
cols = sheet1.max_row
i = 0
for ra in range(1, cols + 1):
    data = sheet1._get_cell(ra, 4).value  # fourth row of the sheet
    try:
        data = data.lower()
        i += 1
    except:
        print("unsuccessful", data)
        print(i)
    rows.append(data)
print(i)
print(rows)

########################################################################################################################


########################################################################################################################

# contains data from dice
workbook2 = load_workbook(filename="QA 8 feb no duplicate.xlsx")
sheet2 = workbook2.active
max_row = sheet2.max_row

########################################################################################################################


########################################################################################################################

# check existing data in collected and blacklist companies
z = 0
i = 1
for i in range(1, max_row):
    try:
        # for i in range(1, max_row + 1):
        # try:
        #     sheet1.delete_rows(position + 1)
        #     rows.pop(position)
        #     workbook1.save(filename="data/August-30/apollo.xlsx")
        # except:
        #     pass
        first = []
        position = ""
        company_name = sheet2._get_cell(i, 2).value  # 2nd row of the sheet
        next_company_name = sheet2._get_cell(i + 1, 2).value
        for item in range(1, sheet2.max_column + 1):
            first.append(sheet2._get_cell(i + 1, item).value)
        print("part 1", first)
        lower_name = company_name.lower()
        if lower_name in rows and lower_name not in blacklist:
            position = rows.index(lower_name)
            for total in range(1, 20):
                company = sheet1._get_cell(position + 1, 4).value
                if company.lower() == company_name.lower():
                    # for item2 in range(1, sheet1.max_column + 1):
                    second = []
                    for item2 in range(1, sheet1.max_column + 1):
                        second.append(sheet1._get_cell(position + 1, item2).value)
                    print("part 2", second)
                    position += 1
                    third = first + second
                    print("Combined", third)
                    third_arr.append(third)
            z += 1
            print(i, z)
            third_arr.append([" "])
            third_arr.append([" "])
            third_arr.append([" "])
        else:
            if lower_name not in blacklist:
                z += 1
                print("not found", lower_name, first)
                fourth_arr.append(first)
                print(z)
                i += 1
                # if not position:
                #     # pass
                #     # Add not collected data in a separate sheet
                #     workbook4 = Workbook()
                #     sheet4 = workbook4.active
                #     sheet4.append(first)
                #     workbook4.save(filename=f"TotallyNew.xlsx")
    except:
        pass

########################################################################################################################


########################################################################################################################
# """Add already present in a sheet"""

# workbook = Workbook()
# sheet = workbook.active
# for i in range(0, len(third_arr)):
#     sheet.append(third_arr[i])
# print("Finished adding data to the sheet")
# workbook.save(filename=f"data_{now}.xlsx")

with open(f"data_{now}.csv", "w", encoding='UTF8', newline='') as n:
    writer = csv.writer(n)
    writer.writerows(third_arr)

########################################################################################################################
# """ Add new and not collected data in a separate sheet to be collected again """

# workbook5 = Workbook()
# sheet5 = workbook5.active
# for i in range(0, len(fourth_arr)):
#     sheet5.append(fourth_arr[i])
# print("Finished adding data to the sheet")
# workbook5.save(filename=f"Not Found_{now}.xlsx")
header = ['Position', 'Company Name', 'Location']
with open(f"Not Found_{now}.csv", "w", encoding='UTF8', newline='') as p:
    writer = csv.writer(p)
    writer.writerow(header)
    writer.writerows(fourth_arr)

########################################################################################################################
