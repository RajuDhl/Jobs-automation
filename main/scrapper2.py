import datetime
import logging
import os
from openpyxl import load_workbook, Workbook
from linkedin_jobs_scraper import LinkedinScraper
from linkedin_jobs_scraper.events import Events, EventData
from linkedin_jobs_scraper.query import Query, QueryOptions, QueryFilters
from linkedin_jobs_scraper.filters import RelevanceFilters, TimeFilters, TypeFilters, \
    ExperienceLevelFilters, RemoteFilters


def scrap():
    # Change root logger level (default is WARN)
    logging.basicConfig(level=logging.INFO)

    # Create a blank xlsx workbook
    data = 'data'
    path = datetime.datetime.now().strftime("%B-%d")
    now = datetime.datetime.now().strftime("%B-%I-%M-%S")
    # print(os.path.isdir(path))
    if os.path.isdir(f'data/{path}'):
        print("dir exists")
    else:
        try:
            os.mkdir(f'{data}')
        except: pass
        os.mkdir(f'data/{path}')
    # writer = pd.ExcelWriter(f'{now}.xlsx', engine='xlsxwriter')
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['query', 'location', 'job_id', 'job_index', 'link',
                  'apply_link', 'title', 'company', 'place', 'date', 'seniority_level', 'job_function',
                  'employment_type', 'industry', 'description',
                  'description_html'])

    workbook.save(filename=f"data/{path}/{now}.xlsx")

    def on_data(data: EventData):
        workbook2 = load_workbook(filename=f"data/{path}/{now}.xlsx")
        sheet2 = workbook2.active
        sheet2.append([data.query, data.location, data.job_id, data.job_index, data.link,
                       data.apply_link, data.title, data.company, data.place, data.date,
                       data.seniority_level, data.job_function, data.employment_type,
                       data.industries, data.description, data.description_html])
        workbook2.save(filename=f"data/{path}/{now}.xlsx")

        print('[ON_DATA]', type(data), data.title, data.company, data.date, data.link, len(data.description))

    def on_error(error):
        print('[ON_ERROR]', error)

    def on_end():
        print('[ON_END]')

    scraper = LinkedinScraper(
        chrome_executable_path="chromedriver.exe",  # Custom Chrome executable path (e.g. /foo/bar/bin/chromedriver)
        chrome_options=None,  # Custom Chrome options here
        headless=True,  # Overrides headless mode only if chrome_options is None
        max_workers=1,
        # How many threads will be spawned to run queries concurrently (one Chrome driver for each thread)
        slow_mo=10,  # Slow down the scraper to avoid 'Too many requests (429)' errors
    )

    # Add event listeners
    scraper.on(Events.DATA, on_data)
    scraper.on(Events.ERROR, on_error)
    scraper.on(Events.END, on_end)

    queries = [
        Query(
            query=f'QA',
            options=QueryOptions(
                locations=['Dallas-Fort Worth Metroplex'],
                optimize=True,  # Blocks requests for resources like images and stylesheet
                limit=1000,  # Limit the number of jobs to scrape
                filters=QueryFilters(
                    # Filter by companies
                    relevance=RelevanceFilters.RECENT,
                    time=TimeFilters.MONTH,
                    type=[TypeFilters.FULL_TIME],
                    experience=[ExperienceLevelFilters.INTERNSHIP, ExperienceLevelFilters.ENTRY_LEVEL,
                                ExperienceLevelFilters.ASSOCIATE],
                )
            )
        )
        # Query(
        #     query='Engineer',
        #     options=QueryOptions(
        #         locations=['United States'],
        #         optimize=False,
        #         limit=5,
        #         filters=QueryFilters(
        #             company_jobs_url='https://www.linkedin.com/jobs/search/?f_C=1441%2C17876832%2C791962%2C2374003%2C18950635%2C16140%2C10440912&geoId=92000000',
        #             # Filter by companies
        #             relevance=RelevanceFilters.RECENT,
        #             time=TimeFilters.MONTH,
        #             type=[TypeFilters.FULL_TIME, TypeFilters.INTERNSHIP],
        #             experience=None,
        #         )
        #     )
        # ),
    ]

    scraper.run(queries)
