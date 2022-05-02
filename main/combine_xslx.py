import datetime
import time

from openpyxl import load_workbook, Workbook
import csv

third_arr = []  # Combined data
fourth_arr = []  # unfounded data (people not found for the companies)

day = datetime.datetime.now().strftime("%B-%d")
now = datetime.datetime.now().strftime("%B-%I-%M-%S")


def blacklist():
    workbook99 = load_workbook(filename='files/Blacklist.xlsx')
    workbook98 = load_workbook(filename='files/QA/Processed 2feb.xlsx')  # Sheet linking to already collected data
    sheet99 = workbook99.active
    sheet98 = workbook98.active
    blacklist = []
    column = sheet99.max_row
    column2 = sheet98.max_row
    for col in range(1, column + 1):
        try:
            valu = sheet99._get_cell(col, 1).value
            valu = valu.lower()
            blacklist.append(valu)
        except:
            pass
    for col2 in range(1, column2 + 1):
        try:
            valuz = sheet98._get_cell(col2, 1).value
            valuz = valuz.lower()
            blacklist.append(valuz)
        except:
            pass
    return blacklist


def main():
    workbook1 = load_workbook(filename="All People.xlsx")
    sheet1 = workbook1.active

    workbook2 = load_workbook(filename="QA 8 feb no duplicate.xlsx")
    sheet2 = workbook2.active

    rows = []

    def apollo():
        for ra in range(1, sheet1.max_row + 1):
            data = sheet1._get_cell(ra, 4).value.lower()  # fourth row of the sheet
            rows.append(data)
        return rows

    for i in range(1, sheet2.max_row + 1):
        try:
            first = []
            company_name = sheet2._get_cell(i, 2).value.lower()
            for item in range(1, sheet2.max_column + 1):
                first.append(sheet2._get_cell(i + 1, item).value)
                

        except: pass

